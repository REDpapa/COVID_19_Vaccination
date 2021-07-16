
from bs4 import BeautifulSoup
import requests

url = "https://kantei.go.jp/jp/headline/kansensho/vaccine.html"
r = requests.get(url)

# 文字化けが酷いのでエンコードします。
r.encoding = r.apparent_encoding

# 引数に'lxml'を指定(なぜ必要かは、忘れちゃったww許して!!)
soup = BeautifulSoup(r.text, 'lxml')

# soupの中身を見たいなら下記をコメントoffして!
# print(soup)

# soup.select("p.aly_tx_center > a")でcssのクラス指定後にaタグを取得
# NOTE:例としてリストで吐き出さるので[0]を指定して開いてます。
#　　　ここを工夫して取れるといけるかもです!!
a_text = soup.select("p.aly_tx_center > a")[0]
link = a_text.get("href")
xlsx_link = f'https://kantei.go.jp{link}'
print(xlsx_link)

"""
<p class="aly_tx_center">
<a href="/jp/content/KOREI-vaccination_data2.xlsx" target="_blank">Excel<img alt="Excelファイルを開く" src="/jp/content/exel_icon.gif">&nbsp;</a>
</p>
"""
root_url = 'https://kantei.go.jp'
xl_url=root_url +link

r = requests.get(xl_url)

# 文字化けが酷いのでエンコードします。
r.encoding = r.apparent_encoding

saveFilePath = "data.xlsx"
with open(saveFilePath, 'wb') as saveFile:
        saveFile.write(r.content)

from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference, Series

wb = load_workbook("data.xlsx")
ws = wb.worksheets[0]

rng1 = ws["A6:G6"]



values = Reference(ws, min_col=4, min_row=6, max_col=4, max_row=10)
x = Reference(ws, min_col=1, min_row=6, max_col=1, max_row=10)
chart = BarChart()
chart.add_data(values)
chart.set_categories(x)
ws.add_chart(chart, "J10")

wb.save("SampleChart.xlsx")
