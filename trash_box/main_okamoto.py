"""
pythonファイルとして実行してみる
"""

# 外部ライブラリーのインポート
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.chart import BarChart,Reference


# ライブラリーのインポート
from datetime import datetime as dt
import requests


# アクセスするULRを指定する
url = "https://kantei.go.jp/jp/headline/kansensho/vaccine.html"
r = requests.get(url)

# 文字化けが酷いのでエンコードします。
r.encoding = r.apparent_encoding

# 引数に'lxml'を指定(なぜ必要かは、忘れちゃったww許して!!)
soup = BeautifulSoup(r.text, 'lxml')

# soup.select("p.aly_tx_center > a")でcssのクラス指定後にaタグを取得
# NOTE:例としてリストで吐き出さるので[0]を指定して開いてます。
#　　　ここを工夫して取れるといけるかもです!!
a_text = soup.select("p.aly_tx_center > a")[0]
link = a_text.get("href")
xlsx_link = f'https://kantei.go.jp{link}'
print(xlsx_link)

# web scraping でお世話になっている requests モジュール、使います!
import requests
# papa サン提供の url を get します!
response = requests.get(xlsx_link)
# response.content に get したファイルが入っているので、保存します!
# NOTE: w は「ファイルに書き込む」、という意味ですよー。
#       b は「そのファイルはバイナリです」、という意味ですよー。
#       xlsx ファイルはバイナリで書いてあるので、そうしないといけないんです。
with open('input.xlsx', 'wb') as saveFile:
    saveFile.write(response.content)

# webサイトから読み出したexcelファイルを読み出してくる。
wb = openpyxl.load_workbook(filename='input.xlsx')
# シートの先頭を指定
sheet_name = wb.sheetnames[0]
ws = wb[sheet_name]
# エクセル表の最終のセルの位置を取得する。
# NOTE: 取得したエクセルファイルの表形式の部分は最終行の'-3'の位置
ws_max_row = ws.max_row - 3

# 7行目のA〜Kまでのデータを取得
# NOTE: 変数rengはタプル型で出力される。
remgs = ws[f'A7:K{ws_max_row}']

# リストを定義
cell_a = []
cell_other = []

# 1行ずつ値を取得していく。
for remg_list in remgs:
    vals = []
    vals = [c.value for c in remg_list]
    a_cell_date_data = dt.strftime(vals[0], '%Y/%m/%d')
    cell_other_data = vals[3:11]
    # cellデータとするため、appendにて追加していく。
    cell_a.append(a_cell_date_data)
    cell_other.append(cell_other_data)

# 新しいシートを作成
new_ws = wb.create_sheet('edit_sheet')

# 項目名を入力していく
new_ws.merge_cells('A1:A3')
new_ws['A1'] = '接種日'
new_ws.merge_cells('B1:E1')
new_ws['B1'] = 'すべて'
new_ws.merge_cells('F1:I1')
new_ws['F1'] = '高齢者'
new_ws.merge_cells('B2:C2')
new_ws.merge_cells('F2:G2')
new_ws['B2'] = new_ws['F2'] = '内1回目'
new_ws.merge_cells('D2:E2')
new_ws.merge_cells('H2:I2')
new_ws['D2'] = new_ws['H2'] = '内2回目'
new_ws['B3'] = new_ws['D3'] = new_ws['F3'] = new_ws['H3'] = 'ファイザー社'
new_ws['C3'] = new_ws['E3'] = new_ws['G3'] = new_ws['I3'] = '武田/モデルナ社'


# cellデータを１つにするためにcell_aを先頭に追加する。
for i in range(len(cell_other)):
    cell_other[i].insert(0,cell_a[i])

# cellの最終行と接種データの総数を加算
cell_max_range = new_ws.max_row + len(cell_other)

# 書き込むセルの範囲を算出。
writing_range = f'A4:J{cell_max_range}'
cell_writing_range = new_ws[writing_range]

# 書き込み範囲のセルにデータを書き込む
for _i in range(len(cell_other)):
    for cell,val in zip(cell_writing_range[_i],cell_other[_i]):
        cell.value=val

# グラフの描画
data = Reference(
    new_ws,
    min_col=2,
    min_row=4,
    max_col=5,
    max_row=new_ws.max_row,
)

x = Reference(
    new_ws,
    min_col=1,
    min_row=4,
    max_col=1,
    max_row=new_ws.max_row,
)

# 新しいシートを作成
graph_ws = wb.create_sheet('graph_sheet')

chart = BarChart()
chart.add_data(data)
chart.set_categories(x)
graph_ws.add_chart(chart, "A1")

# 現在ファイルとして保存
wb.save('input.xlsx')
