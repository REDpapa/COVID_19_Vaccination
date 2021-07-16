from bs4 import BeautifulSoup
import requests

url = "https://kantei.go.jp/jp/headline/kansensho/vaccine.html"
r = requests.get(url)

# 文字化けが酷いのでエンコードします。
r.encoding = r.apparent_encoding

# 引数に'lxml'を指定(なぜ必要かは、忘れちゃったww許して!!)
soup = BeautifulSoup(r.text, 'lxml')

# souの中身を見たいなら下記をコメントoffして!
#print(soup)

# http://soup.select("p.aly_tx_center > a")でcssのクラス指定後にaタグを取得
# NOTE:例としてリストで吐き出さるので[0]を指定して開いてます。
#　　　ここを工夫して取れるといけるかもです!!
a_text = soup.select("p.aly_tx_center > a")[0]
link = a_text.get("href")
#print(link)

root_url = 'https://kantei.go.jp'
xl_url=root_url +link

r = requests.get(xl_url)
saveFilePath = "wakutinn2.xlsx"
with open(saveFilePath, 'wb') as saveFile:
        saveFile.write(r.content)

import openpyxl
import datetime
import locale

#today = datetime.date.today()
#locale.setlocale(locale.LC_TIME,'ja_JP.UTF-8')
#print(locale.getlocale(locale.LC_TIME))
#print(today.strftime('%A'))

wb = openpyxl.load_workbook("wakutinn2.xlsx")
ws = wb.worksheets[0]
remg = ws['A7:K7']
vals = [c.value for c in remg[0]]
#print(vals)
wb2 = openpyxl.load_workbook("LT3.xlsx")
ws2 = wb2.worksheets[0]
row = ws2.max_row+1
remg2 = ws2[f'A{row}:K{row}']
for sell,val in zip(remg2[0],vals):
    sell.value=val

wb2.save("LT3.xlsx")
